import os
import pandas as pd
import json
import logging
import openpyxl
import re
from version import __version__
from core.services.logger_service import LoggerService

class ExportService:
    def __init__(self):
        self.logger = LoggerService()

    def export(self, df, filepath_or_format, format_type=None, sheet_name=None, key_col=None):
        """
        Exporta o DataFrame para o formato desejado.
        Suporta MODO CIRÚRGICO (Preservação) para Excel.
        """
        if format_type and not filepath_or_format.endswith(format_type):
            filepath = filepath_or_format + format_type
        else:
            filepath = filepath_or_format
            
        ext = os.path.splitext(filepath)[1].lower()
        
        if ext == '.csv':
            df.to_csv(filepath, index=False, sep=';', encoding='utf-8-sig')
            return True
        
        elif ext == '.xlsx':
            target_sheet = sheet_name or "Sheet1"
            # Se o arquivo já existe e temos a chave, usamos Injeção Cirúrgica
            if os.path.exists(filepath) and key_col:
                return self.export_surgical(df, filepath, target_sheet, key_col)
            else:
                df.to_excel(filepath, sheet_name=target_sheet, index=False)
                return True
            
        return False

    def export_surgical(self, df, filepath, sheet_name, key_col):
        """
        Injeta dados em um Excel existente linha a linha, preservando TUDO (Fórmulas/Estilos).
        """
        # Normalização ULTRA-ROBUSTA
        def clean_key(v):
            if v is None: return ""
            s = str(v).strip().upper()
            s = re.sub(r'\.0$', '', s)
            s = re.sub(r'[.,\-/]', '', s)
            return s.lstrip('0')

        try:
            wb = openpyxl.load_workbook(filepath)
            if sheet_name not in wb.sheetnames:
                self.logger.warning(f"Aba '{sheet_name}' não encontrada. Usando aba ativa.")
                ws = wb.active # Fallback para a primeira aba detectada
            else:
                ws = wb[sheet_name]

            # Mapear cabeçalhos
            headers = [str(ws.cell(row=1, column=c).value).strip() for c in range(1, ws.max_column + 1)]
            
            try:
                key_idx_ws = headers.index(key_col) + 1
            except ValueError:
                self.logger.error(f"Chave '{key_col}' não encontrada no Excel.")
                return False

            map_cols = {}
            for col in df.columns:
                if col in headers and col != key_col:
                    map_cols[col] = headers.index(col) + 1
            
            df_lookup = df.copy()
            df_lookup['_K'] = df_lookup[key_col].apply(clean_key)
            df_lookup = df_lookup.drop_duplicates(subset=['_K'], keep='first')
            lookup = df_lookup.set_index('_K').to_dict('index')

            updated_count = 0
            for r in range(2, ws.max_row + 1):
                raw_val = ws.cell(row=r, column=key_idx_ws).value
                if raw_val is None: continue
                
                clean_k = clean_key(raw_val)
                if clean_k in lookup:
                    row_data = lookup[clean_k]
                    for col_name, col_idx in map_cols.items():
                        new_val = row_data.get(col_name)
                        if pd.notna(new_val):
                            ws.cell(row=r, column=col_idx, value=new_val)
                            updated_count += 1

            wb.save(filepath)
            self.logger.info(f"Sucesso: {updated_count} células injetadas em '{os.path.basename(filepath)}'")
            return True
        except Exception as e:
            self.logger.exception(f"Erro na exportação cirúrgica: {e}")
            return False

# --- Declaração de Versão do Módulo (Genaja Version Hook) ---
from version_hook import declare as _vdeclare
_vdeclare(__name__, __version__, "Serviço de exportação enterprise com injeção cirúrgica de dados")
